import pandas as pd
import xlsxwriter
import openpyxl
import requests
import urllib.request
import json
import certifi
import ssl
from methods import *

# Henter inn input-filen
excelInputWorkBook = openpyxl.load_workbook('SS24 Main Ganni import.xlsx', data_only=True)
excelInputWorkSheet = excelInputWorkBook.worksheets[0]

# Oppretter en excel-fil som skal brukes som look-up mot API'et til Front
excelCheckUp = xlsxwriter.Workbook("CheckUp.xlsx")
checkUpWorkSheet = excelCheckUp.add_worksheet()


#Legger inn overskrifter i excel-filen som skal brukes som look-up mot Front sitt API og som skal bli output-fil til slutt

checkUpWorkSheet.write('A1', 'EAN')
checkUpWorkSheet.write('B1', 'New Product / Product ID')
checkUpWorkSheet.write('C1', 'In price changed?')
checkUpWorkSheet.write('D1', 'InPriceNew')
checkUpWorkSheet.write('E1', 'InPriceBefore')
checkUpWorkSheet.write('F1', 'Out price change?')
checkUpWorkSheet.write('G1', 'OutPriceNew')
checkUpWorkSheet.write('H1', 'OutPriceBefore')
checkUpWorkSheet.write('I1', 'Season changed?')
checkUpWorkSheet.write('J1', 'SeasonNew')
checkUpWorkSheet.write('K1', 'SeasonBefore')

# Lager variabler for rad og kolonne for EAN til checkup-filen
eanRow = 1
eanColumn = 0

# Lager variabler for rad og kolonne for innpris før til checkup-filen
inPriceRow = 1
inPriceColumn = 3

# Lager variabler for rad og kolonne for utpris før til checkup-filen
outPriceRow = 1
outPriceColumn = 6

seasonRow = 1
seasonColumn = 9

"""
    workbook = xlsxwriter.Workbook('Example.xlsx')
    worksheet = workbook.add_worksheet()
"""

# Itererer gjennom den filen som kommer inn som input for å finne EAN, innpris og utpris og legger dette inn i riktig kolonne i lookup-filen

for column in excelInputWorkSheet.iter_cols():
    column_name = column[0].value
    if column_name == "EAN":
        for cell in column:
            if cell.value == "EAN":
                continue

            EAN = str(cell.value)
            if len(EAN) < 13 and EAN != 'None':
                EAN = "0" + EAN
                
            if EAN != 'None':
                checkUpWorkSheet.write(eanRow, eanColumn, EAN)
            eanRow += 1

    if column_name == "InPrice":
        for cell in column:
            if cell.value == "InPrice":
                continue

            checkUpWorkSheet.write(inPriceRow, inPriceColumn, cell.value)
            inPriceRow += 1

    if column_name == "OutPrice":
        for cell in column:
            if cell.value == "OutPrice":
                continue
            checkUpWorkSheet.write(outPriceRow, outPriceColumn, cell.value)
            outPriceRow += 1 

    if column_name == "Season":
        for cell in column:
            if cell.value == "Season":
                continue

            season = str(cell.value)
            if season != 'None':
                checkUpWorkSheet.write(seasonRow, seasonColumn, season)
            seasonRow += 1 


excelCheckUp.close()
print("Excel er laget")

# Neste her er å bruke API'et til Front for å
# 1) Finne ut om produktet ligger i Front fra før
# 2) Hente ut tidligere innpris og utpris på produktet
# 3) Legge inn om utpris (i første omgang) er endret
# 4) Legge inn ny innpris i lookup-filen 
# 5) Legge inn ny utpris i lookup-filen
# Etterhvert her vil det også være aktuelt å finne ut hvilke butikker som har dette produktet på lager og så eventuelt legge inn det i filen også

#response = requests.get("https://frontsystemsapis.frontsystems.no/restapi/V2/api/Pricelist/gtin/{gtin}")
#print(response.status_code)

# Laster inn checkup-filen som skal sjekkes mot API'et i Front
workbookOutput = openpyxl.load_workbook("CheckUp.xlsx")
wb = workbookOutput.active

# Variables som ikke er brukt enda, prøver å finne en bedre løsning på å få satt riktig info i riktig kolonne i for-løkken under istedenfor å hard-kode kolonneindeks
columnCount = wb.max_column
eanColumn = columnCount - (columnCount-1)
infoColumn = columnCount - (columnCount-2)
inPriceChangeColumn = columnCount - (columnCount-3)

cellCounter = 1
rowCounter = 2

for row in wb.iter_rows(min_row=2, max_col=columnCount):
    print(rowCounter)
    ProductFromFront = []
    for cell in row:
        if cellCounter == 1:
            EAN = str(cell.value)
            ProductFromFront = json.load(findPriceFromGtin(EAN))
            if ProductFromFront == []:
                wb.cell(rowCounter, 2, "New product")
                rowCounter += 1
                break
            else:
                wb.cell(rowCounter, 2, ProductFromFront[0]['productid'])

        if cellCounter == 4:
            InPriceNew = cell.value
            InPriceOld = ProductFromFront[0]['cost']
            if InPriceNew != InPriceOld:
                wb.cell(rowCounter, 3, "Yes")
            else:
                wb.cell(rowCounter, 3, "No")

            wb.cell(rowCounter, 5, InPriceOld)

        if cellCounter == 7:
            OutPriceNew = cell.value
            OutPriceOld = ProductFromFront[0]['price']
            if OutPriceNew != OutPriceOld:
                wb.cell(rowCounter, 6, "Yes")
            else:
                wb.cell(rowCounter, 6, "No")

            wb.cell(rowCounter, 8, OutPriceOld)
    
        if cellCounter == 10:
            SeasonNew = cell.value
            SeasonOld = ProductFromFront[0]['season']
            if SeasonNew != SeasonOld:
                wb.cell(rowCounter, 9, "Yes")
            else:
                wb.cell(rowCounter, 9, "No")

            wb.cell(rowCounter, 11, SeasonOld)
            
            cellCounter = 1
            rowCounter += 1
            break

        cellCounter += 1

         

workbookOutput.save("FinalOutput.xlsx")

        # product = findPriceFromGtin(str(cell.value))



"""
try:
    url = "https://frontsystemsapis.frontsystems.no/restapi/V2/api/Pricelist/gtin/3260645641410"

    hdr ={
    # Request headers
    'x-api-key': 'rSzTE8K.4PYkiCCRpFS89sfyhD6QFvSKJrvsc8Gi',
    'Cache-Control': 'no-cache',
    'Ocp-Apim-Subscription-Key': 'ba812cd412304d1daa83c558113acacc',
    }

    gcontext = ssl.create_default_context(cafile=certifi.where())

    req = urllib.request.Request(url, headers=hdr)

    req.get_method = lambda: 'GET'
    response = urllib.request.urlopen(req, context=gcontext)
    print(response.getcode())
    print(response.read())
except Exception as e:
    print(e)
"""



"""
for i in range(1, test.max_row+1):
    cell_obj = test.cell(row=i, column = i)
    print(cell_obj.value)


for item in dataframe1 :
    worksheet.write(row, column, item)
    column += 1
   
workbook.close()
"""
