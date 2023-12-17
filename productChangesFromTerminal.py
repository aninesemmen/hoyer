#!/usr/bin/env python3

import pandas as pd
import xlsxwriter
import argparse
import openpyxl
import os
from methods import *

# Henter input fra argument som er med når scriptet kjøres
parser = argparse.ArgumentParser(description="Filnavn for inputfil")
parser.add_argument("filename", help="Skriv inn filnavnet til input-filen som første argument")
parser.add_argument("outputDirectory", help="Spesifiser filsti til hvor du vil at output-fil skal lagres")
args = parser.parse_args()

excelInputFileName = args.filename

# Henter inn import-fil, altså den filen som skal sjekkes, lagrer først navnet på filen i en variabel her, dette må da endres manuelt avhengig av hvilken fil som skal sjekkes 
# excelInputFileName = 'SS24 Main_AW24 Calvin Klein import.xlsx'

excelInputWorkBook = openpyxl.load_workbook(excelInputFileName, data_only=True)
excelInputWorkSheet = excelInputWorkBook.worksheets[0]

# Sletter først en eventuelt eksisterende lookup-fil slik at det ikke blir liggende gammel data i filen
# Bruker xlsxwriter for å opprette en excel-fil med standardisert format som skal brukes som look-up mot API'et til Front og danner grunnlaget for output-filen
# lookupFileName-variabelen brukes til å sette navnet på excel-filen, grunnen til at den er laget som variabel er fordi navnet brukes flere steder og i tilfelle dette senere skal endres trenger det da bare endres et sted
lookUpFileName = 'LookUp.xlsx'
if os.path.isfile(lookUpFileName):
    os.remove(lookUpFileName)

excelCheckUp = xlsxwriter.Workbook(lookUpFileName)
checkUpWorkSheet = excelCheckUp.add_worksheet()

columnNames = ['EAN','New Product / Product ID', 'Product name', 
               'In price changed?', 'InPriceNew', 'InPriceBefore', 
               'Out price changed?', 'OutPriceNew', 'OutPriceBefore', 'Season changed?', 'SeasonNew', 'SeasonBefore']

# Legger inn kolonneoverskrifter i lookup-filen, dette danner også grunnlaget for endelig output-fil
# Her er målet å få skrevet litt mer kompakt kode med ikke hardkodede kolonnereferanser (A1, B1 osv). Dette fordi denne løsningen er lite praktisk og krever manuell endring hvis det skal legges til flere kolonner senere.
checkUpWorkSheet.write('A1', 'EAN')
checkUpWorkSheet.write('B1', 'New Product / Product ID')
checkUpWorkSheet.write('C1', 'Product name')
checkUpWorkSheet.write('D1', 'In price changed?')
checkUpWorkSheet.write('E1', 'InPriceNew')
checkUpWorkSheet.write('F1', 'InPriceBefore')
checkUpWorkSheet.write('G1', 'Out price change?')
checkUpWorkSheet.write('H1', 'OutPriceNew')
checkUpWorkSheet.write('I1', 'OutPriceBefore')
checkUpWorkSheet.write('J1', 'Season changed?')
checkUpWorkSheet.write('K1', 'SeasonNew')
checkUpWorkSheet.write('L1', 'SeasonBefore')

# Lager variabler for rad og kolonne for EAN til checkup-filen
eanRow = 1
eanColumn = 0

# Lager variabler for rad og kolonne for innpris før til checkup-filen
inPriceRow = 1
inPriceColumn = 4

# Lager variabler for rad og kolonne for utpris før til checkup-filen
outPriceRow = 1
outPriceColumn = 7

seasonRow = 1
seasonColumn = 10

# Itererer gjennom import-filen for å finne EAN, nyeste innpris og nyeste utpris og legger dette inn i riktig kolonne i lookup-filen
# Her brukes kolonnenavnet for å avgjøre hva som skal gjøres videre med verdiene, dette fordi verdier skal settes i ulike kolonner avhengig av hvilken verdi vi ser på
# Eksempelvis på EAN må det sjekkes om dette har 13 siffer, hvis ikke må det legges til en 0 på starten for å kunne få sammenlignet med Front-verdier
# Continue-statements brukes for å hoppe over overskriftene

for column in excelInputWorkSheet.iter_cols():
    column_name = column[0].value
    if column_name == "EAN":
        for cell in column:
            if cell.value == "EAN":
                continue

            EAN = str(cell.value)
            if len(EAN) < 13 and EAN != 'None':
                EAN = "0" + EAN
                
            if EAN != 'None':
                checkUpWorkSheet.write(eanRow, eanColumn, EAN)

            eanRow += 1

    if column_name == "InPrice":
        for cell in column:
            if cell.value == "InPrice":
                continue

            checkUpWorkSheet.write(inPriceRow, inPriceColumn, cell.value)
            inPriceRow += 1

    if column_name == "OutPrice":
        for cell in column:
            if cell.value == "OutPrice":
                continue
            checkUpWorkSheet.write(outPriceRow, outPriceColumn, cell.value)
            outPriceRow += 1 

    if column_name == "Season":
        for cell in column:
            if cell.value == "Season":
                continue

            season = str(cell.value)
            if season != 'None':
                checkUpWorkSheet.write(seasonRow, seasonColumn, season)
            seasonRow += 1 


excelCheckUp.close()
print("LookUp-fil er laget")

# Her brukes panda for å lese inn lookup-filen
readExcel = pd.read_excel(lookUpFileName,sheet_name=0, dtype={"EAN":str})

# Lages en liste over EANs fra lookup-filen slik at denne kan sendes inn i metoden som jobber mot Front sitt API og få hentet ut eventuelle produkter som allerede ligger inne
listOfEans = readExcel['EAN'].tolist()

# Bruker metoden som jobber mot API'et til Front og henter ut eksisterende produkter, input parameter er listen med EANs som ble laget over
productList = findPriceFromGtins({"gtins":listOfEans})

existingGtins = []

for item in productList:
    for gtin in item['productSizes']:
        existingGtins.append(gtin['gtin'])

# Brukes openpyxl på look-up filen for å få tilgang til å iterere gjennom (mulig dette også kan gjøres med panda så man slipper å bruke mange pakker, har ikke testet dette)
workbookOutput = openpyxl.load_workbook(lookUpFileName)
wb = workbookOutput.active

columnCheck = workbookOutput.worksheets[0]

# Setter kolonne-teller til det totale antallet kolonner som finnes i look-up-filen
columnCount = wb.max_column

# Setter en celle-teller slik at det er mulig å håndtere de forskjellige cellene på ulike måter når man itererer gjennom hver celle i hver rad lenger ned
cellCounter = 1

# Setter en rad-teller slik at man får satt riktig verdi på riktig rad
rowCounter = 2

# Lager egne variabler for kolonnonumre som skal brukes i col-loopen under for å få satt verdiene på riktig sted uten å måtte hardkode kolonnenumrene
eanColumn = 0
newProductColumn = 0
productNameColumn = 0
inPriceNewColumn = 0
outPriceNewColumn = 0
seasonNewColumn = 0

for col in columnCheck.iter_cols(min_row=1, max_col=columnCount,max_row=1):
    column_name = col[0].value
    column_number = col[0].column
    match column_name:
        case "EAN":
            eanColumn = column_number
        case "New Product / Product ID":
            newProductColumn = column_number
        case "Product name":
            productNameColumn = column_number
        case "InPriceNew":
            inPriceNewColumn = column_number
        case "OutPriceNew":
            outPriceNewColumn = column_number
        case "SeasonNew":
            seasonNewColumn = column_number

# Iterer gjennom look-up filen rad for rad og bruker listen med allerede eksiterende produkter som ble laget lenger opp til å sette inn om det er nytt produkt, evt ny pris, ny sesong etc
# Må her finne en bedre løsning på celle-telleren fordi denne må endres manuelt hvis det skjer noen som helst endringer i filen, for eksempel hvis det legges inn en ny kolonne lenger opp. Lite praktisk.
for row in wb.iter_rows(min_row=2, max_col=columnCount):
    productActual = []
    for cell in row:
        if cellCounter == eanColumn:
            EAN = str(cell.value)
            if EAN not in existingGtins:
                wb.cell(rowCounter, newProductColumn, "New product")
                rowCounter += 1
                break
            else:
                for product in productList:
                    for gtin in product['productSizes']:
                        if gtin['gtin'] == EAN:
                            productActual = product
                wb.cell(rowCounter, newProductColumn, productActual['productid'])

        if cellCounter == productNameColumn:
            wb.cell(rowCounter, productNameColumn, productActual['name'])
        
        if cellCounter == inPriceNewColumn:
            InPriceNew = cell.value
            InPriceOld = productActual['cost']
            if InPriceNew != InPriceOld:
                wb.cell(rowCounter, inPriceNewColumn-1, "Yes")
            else:
                wb.cell(rowCounter, inPriceNewColumn-1, "No")

            wb.cell(rowCounter, inPriceNewColumn+1, InPriceOld)

        if cellCounter == outPriceNewColumn:
            OutPriceNew = cell.value
            OutPriceOld = productActual['price']
            if OutPriceNew != OutPriceOld:
                wb.cell(rowCounter, outPriceNewColumn-1, "Yes")
            else:
                wb.cell(rowCounter, outPriceNewColumn-1, "No")

            wb.cell(rowCounter, outPriceNewColumn+1, OutPriceOld)
    
        if cellCounter == seasonNewColumn:
            SeasonNew = cell.value
            SeasonOld = productActual['season']
            if SeasonNew != SeasonOld:
                wb.cell(rowCounter, seasonNewColumn-1, "Yes")
            else:
                wb.cell(rowCounter, seasonNewColumn-1, "No")

            wb.cell(rowCounter, seasonNewColumn+1, SeasonOld)
            
            cellCounter = 1
            rowCounter += 1
            break

        cellCounter += 1

outputFileName = "FinalOutput_" + excelInputFileName
outputDirectory = args.outputDirectory
if os.path.isfile(outputFileName):
    os.remove(outputFileName)

if os.path.isfile(lookUpFileName):
    os.remove(lookUpFileName)

workbookOutput.save("outputDirectory"+"/"+outputFileName)
workbookOutput.save("OutputFiles/"+outputFileName)
print("Final output er laget")